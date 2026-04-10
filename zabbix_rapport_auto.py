#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script automatise de rapport Zabbix v5 - Your Company
Recommandations IA via Ollama (local), filtrage, categorisation.
"""

import json, urllib.request, ssl, smtplib, os, sys, re
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ============================================================
# CONFIGURATION
# ============================================================
ZABBIX_URL = "https://your-zabbix-server/api_jsonrpc.php"
ZABBIX_USER = os.environ.get("ZABBIX_USER", "rapport-auto")
ZABBIX_PASS = os.environ.get("ZABBIX_PASS", "changeme")

SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.example.com")
SMTP_PORT = 587
SMTP_USER = os.environ.get("SMTP_USER", "your-smtp-user")
SMTP_PASS = os.environ.get("SMTP_PASS", "changeme")
SMTP_FROM = os.environ.get("SMTP_FROM", "Zabbix Alerts <alerts@example.com>")
EMAIL_TO = os.environ.get("EMAIL_TO", "admin@example.com").split(",")

REPORT_DIR = "/opt/zabbix-reports"

# Ollama (IA locale)
OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
OLLAMA_MODEL = "gemma3:1b"

# ============================================================
# FILTRAGE
# ============================================================
EXCLUDED_PATTERNS = [
    r"Ethernet has changed to lower speed",
    r"Operating system description has changed",
    r"GoogleUpdater",
    r"Number of installed packages has been changed",
]
EXCLUDED_SEVERITIES = ["0", "1"]
NETWORK_KEYWORDS = ["aruba", "hp-2530", "switch", "p18ch", "vn53"]
NETWORK_AGENTS = ["2"]

# ============================================================
# SSL - ignorer les certificats auto-signes
# ============================================================
ssl_ctx = ssl.create_default_context()
ssl_ctx.check_hostname = False
ssl_ctx.verify_mode = ssl.CERT_NONE
opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=ssl_ctx))


# ============================================================
# API ZABBIX
# ============================================================
def zabbix_api(method, params, auth=None):
    payload = {"jsonrpc": "2.0", "method": method, "params": params, "id": 1}
    headers = {"Content-Type": "application/json-rpc"}
    if auth is not None and method != "user.login":
        headers["Authorization"] = f"Bearer {auth}"
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(ZABBIX_URL, data=data, headers=headers)
    try:
        resp = opener.open(req, timeout=30)
        result = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        raise Exception(f"Erreur API Zabbix: {e}")
    if "error" in result:
        raise Exception(f"Erreur API Zabbix: {result['error']}")
    return result["result"]


def get_auth_token():
    return zabbix_api("user.login", {"username": ZABBIX_USER, "password": ZABBIX_PASS})


def get_hosts(auth):
    return zabbix_api("host.get", {
        "output": ["hostid", "host", "name", "status"],
        "selectInterfaces": ["ip", "type", "available"],
        "selectGroups": ["name"],
        "sortfield": "name"
    }, auth)


def get_problems(auth):
    return zabbix_api("problem.get", {
        "output": ["eventid", "objectid", "name", "severity", "clock",
                    "r_clock", "acknowledged"],
        "selectTags": "extend",
        "recent": True,
        "sortfield": "eventid",
        "sortorder": "DESC",
        "suppressed": False
    }, auth)


def get_triggers(auth, trigger_ids):
    if not trigger_ids:
        return []
    return zabbix_api("trigger.get", {
        "output": ["triggerid", "description", "priority"],
        "triggerids": trigger_ids,
        "selectHosts": ["host", "name"],
        "expandDescription": True
    }, auth)


def get_host_availability(auth):
    hosts = zabbix_api("host.get", {
        "output": ["hostid"],
        "selectInterfaces": ["available"],
        "filter": {"status": 0}
    }, auth)
    a = u = k = 0
    for h in hosts:
        if h.get("interfaces"):
            v = h["interfaces"][0].get("available", "0")
            if v == "1": a += 1
            elif v == "2": u += 1
            else: k += 1
        else:
            k += 1
    return {"total": len(hosts), "available": a, "unavailable": u, "unknown": k}


# ============================================================
# UTILITAIRES
# ============================================================
def severity_name(sev):
    return {"0": "Non classe", "1": "Information", "2": "Avertissement",
            "3": "Moyen", "4": "Haut", "5": "Desastre"}.get(str(sev), "Inconnu")


def is_excluded(problem):
    if str(problem.get("severity", "0")) in EXCLUDED_SEVERITIES:
        return True
    name = problem.get("name", "") or ""
    for pattern in EXCLUDED_PATTERNS:
        if re.search(pattern, name, re.IGNORECASE):
            return True
    return False


def ask_ollama(prompt):
    """Envoie un prompt a Ollama et retourne la reponse."""
    try:
        payload = {"model": OLLAMA_MODEL, "prompt": prompt, "stream": False,
                   "options": {"temperature": 0.3, "num_predict": 150}}
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(OLLAMA_URL, data=data,
                                     headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        return result.get("response", "").strip()
    except Exception as e:
        return f"Erreur IA: {e}"


def get_recommendation(problem_name, host_name, severity, duration):
    """Genere une recommandation via Ollama."""
    prompt = f"""Donne UNIQUEMENT la recommandation en 2-3 phrases max. Pas d introduction. Inclus les commandes si pertinent. Reponds en francais.

Hote: {host_name}
Probleme: {problem_name}
Severite: {severity}
Duree: {duration}

Recommandation:"""
    return ask_ollama(prompt)


def classify_host(host_name, hosts_data):
    name_lower = (host_name or "").lower()
    for kw in NETWORK_KEYWORDS:
        if kw in name_lower:
            return "Reseau"
    for h in hosts_data:
        if h.get("name", "") == host_name or h.get("host", "") == host_name:
            if h.get("interfaces"):
                if h["interfaces"][0].get("type") in NETWORK_AGENTS:
                    return "Reseau"
    if name_lower.startswith("sx") and not name_lower.startswith("sxssr"):
        return "Poste"
    if name_lower.startswith("imp "):
        return "Peripherique"
    return "Serveur"


# ============================================================
# GENERATION DU RAPPORT EXCEL (AVEC DESIGN)
# ============================================================
def generate_report(hosts, problems, triggers_map, availability):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    now = datetime.now()
    wk_s = now - timedelta(days=now.weekday())
    wk_e = wk_s + timedelta(days=6)

    # Styles
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
    bottom_border = Border(bottom=Side(style='medium', color='1F4E79'))
    title_font = Font(name='Calibri', bold=True, size=18, color='1F4E79')
    subtitle_font = Font(name='Calibri', size=10, italic=True, color='808080')
    section_font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    normal_font = Font(name='Calibri', size=10)
    metric_label_font = Font(name='Calibri', size=9, color='606060')
    alert_font = Font(name='Calibri', size=10, bold=True, color='CC0000')
    reco_font = Font(name='Calibri', size=9, color='1565C0')

    light_gray = PatternFill('solid', fgColor='F8F9FA')
    white_fill = PatternFill('solid', fgColor='FFFFFF')
    green_metric = PatternFill('solid', fgColor='E8F5E9')
    red_metric = PatternFill('solid', fgColor='FFEBEE')
    yellow_metric = PatternFill('solid', fgColor='FFF8E1')
    blue_metric = PatternFill('solid', fgColor='E3F2FD')
    alert_bg = PatternFill('solid', fgColor='FFF0F0')
    dark_header = PatternFill('solid', fgColor='1F4E79')

    sev_fills = {
        "Desastre": PatternFill('solid', fgColor='D32F2F'),
        "Haut": PatternFill('solid', fgColor='E64A19'),
        "Moyen": PatternFill('solid', fgColor='F57C00'),
        "Avertissement": PatternFill('solid', fgColor='FBC02D'),
    }
    sev_fonts = {
        "Desastre": Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
        "Haut": Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
        "Moyen": Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
        "Avertissement": Font(name='Calibri', size=10, bold=True, color='333333'),
    }
    cat_configs = {
        "Serveur": {"fill": PatternFill('solid', fgColor='1565C0'), "icon": "SERVEURS",
                     "row_fill": PatternFill('solid', fgColor='F5F9FF')},
        "Reseau": {"fill": PatternFill('solid', fgColor='E65100'), "icon": "EQUIPEMENTS RESEAU",
                    "row_fill": PatternFill('solid', fgColor='FFF8F0')},
        "Poste": {"fill": PatternFill('solid', fgColor='2E7D32'), "icon": "POSTES DE TRAVAIL",
                   "row_fill": PatternFill('solid', fgColor='F5FFF5')},
        "Peripherique": {"fill": PatternFill('solid', fgColor='6A1B9A'), "icon": "PERIPHERIQUES",
                          "row_fill": PatternFill('solid', fgColor='FCF5FF')},
    }

    # Filtrer et categoriser
    filtered = [p for p in problems if not is_excluded(p)]
    excluded_count = len(problems) - len(filtered)

    categorized = {"Serveur": [], "Reseau": [], "Poste": [], "Peripherique": []}
    total = len(filtered)

    for idx, p in enumerate(filtered):
        ti = triggers_map.get(p.get("objectid", ""), {})
        hosts_list = ti.get("hosts") or []
        hn = hosts_list[0].get("name", "") if hosts_list else ""
        cat = classify_host(hn, hosts)
        if cat not in categorized:
            cat = "Serveur"

        try:
            et = datetime.fromtimestamp(int(p.get("clock", 0)))
        except:
            et = now

        dur = now - et
        total_sec = int(dur.total_seconds())
        d = total_sec // 86400
        h = (total_sec % 86400) // 3600
        m = (total_sec % 3600) // 60
        duration_str = f"{d}j {h}h" if d > 0 else f"{h}h {m}m"

        sev = severity_name(p.get("severity", "0"))

        print(f"  IA [{idx+1}/{total}] {hn}: {p.get('name','')[:50]}...", end=" ", flush=True)
        reco = get_recommendation(p.get("name", ""), hn, sev, duration_str)
        reco = reco.replace("\n", " ").strip()
        if len(reco) > 300:
            reco = reco[:297] + "..."
        print("OK")

        categorized[cat].append({
            "date": et.strftime("%d/%m/%Y %H:%M"), "host": hn, "severity": sev,
            "problem": p.get("name", ""), "duration": duration_str,
            "ack": "Oui" if str(p.get("acknowledged", "0")) == "1" else "Non",
            "sev_code": str(p.get("severity", "0")), "reco": reco,
        })

    sev_count = {"Desastre": 0, "Haut": 0, "Moyen": 0, "Avertissement": 0}
    for cat_list in categorized.values():
        for p in cat_list:
            if p["severity"] in sev_count:
                sev_count[p["severity"]] += 1

    # ===== FEUILLE 1 : Rapport =====
    ws = wb.active
    ws.title = "Rapport Hebdo"
    ws.sheet_properties.tabColor = "1F4E79"
    for col, w in {'A': 14, 'B': 15, 'C': 12, 'D': 35, 'E': 9, 'F': 65}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A2:F2')
    ws['A2'] = 'RAPPORT DE SUPERVISION'
    ws['A2'].font = title_font
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 35

    ws.merge_cells('A3:F3')
    ws['A3'] = (f'Semaine du {wk_s.strftime("%d/%m/%Y")} au {wk_e.strftime("%d/%m/%Y")}  |  '
                f'Your Company  |  {now.strftime("%d/%m/%Y %H:%M")}  |  Recommandations IA (Gemma3)')
    ws['A3'].font = subtitle_font

    for c in range(1, 7):
        ws.cell(row=4, column=c).border = bottom_border

    # Metriques
    row = 6
    metrics = [
        (availability['total'], "Hotes total", blue_metric, '1565C0'),
        (availability['available'], "Disponibles", green_metric, '2E7D32'),
        (availability['unavailable'], "Non disponibles", red_metric, 'C62828'),
        (len(filtered), "Alertes", yellow_metric, 'E65100'),
        (sev_count['Haut'] + sev_count['Desastre'], "Critiques",
         red_metric if sev_count['Haut'] + sev_count['Desastre'] > 0 else green_metric,
         'C62828' if sev_count['Haut'] + sev_count['Desastre'] > 0 else '2E7D32'),
        (excluded_count, "Filtrees", light_gray, '808080'),
    ]
    for i, (val, label, fill, color) in enumerate(metrics):
        col = i + 1
        cell_v = ws.cell(row=row, column=col, value=val)
        cell_v.font = Font(name='Calibri', size=20, bold=True, color=color)
        cell_v.fill = fill
        cell_v.alignment = Alignment(horizontal='center', vertical='center')
        cell_v.border = thin_border
        ws.row_dimensions[row].height = 40
        cell_l = ws.cell(row=row+1, column=col, value=label)
        cell_l.font = metric_label_font
        cell_l.fill = fill
        cell_l.alignment = Alignment(horizontal='center', vertical='center')
        cell_l.border = thin_border

    # Points d'attention
    row = 9
    attention = [f"{p['host']}: {p['problem'][:50]}" for cl in categorized.values()
                 for p in cl if p["severity"] in ("Haut", "Desastre")]
    if attention:
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "POINTS D'ATTENTION"
        ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=10, color='C62828')
        ws[f'A{row}'].fill = alert_bg
        ws[f'A{row}'].border = thin_border
        for item in attention[:5]:
            row += 1
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f'  {item}'
            ws[f'A{row}'].font = Font(name='Calibri', size=10, color='C62828')
            ws[f'A{row}'].fill = alert_bg
            ws[f'A{row}'].border = thin_border

    # Alertes par categorie
    row += 2
    headers = ['Date', 'Hote', 'Severite', 'Probleme', 'Duree', 'Recommandation IA']
    for cat_key in ["Serveur", "Reseau", "Poste", "Peripherique"]:
        items = categorized.get(cat_key, [])
        if not items:
            continue
        conf = cat_configs[cat_key]

        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f'  {conf["icon"]}  ({len(items)})'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = conf["fill"]
        ws[f'A{row}'].alignment = Alignment(vertical='center')
        ws[f'A{row}'].border = thin_border
        ws.row_dimensions[row].height = 28
        row += 1

        for ci, hdr in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=hdr)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='37474F')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1

        sev_order = {"Desastre": 0, "Haut": 1, "Moyen": 2, "Avertissement": 3}
        items.sort(key=lambda x: sev_order.get(x["severity"], 5))

        for idx2, prob in enumerate(items):
            row_fill = conf["row_fill"] if idx2 % 2 == 0 else white_fill
            vals = [prob["date"], prob["host"], prob["severity"],
                    prob["problem"], prob["duration"], prob["reco"]]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = reco_font if ci == 6 else normal_font
                cell.border = thin_border
                cell.fill = row_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            sev = prob["severity"]
            if sev in sev_fills:
                ws.cell(row=row, column=3).fill = sev_fills[sev]
                ws.cell(row=row, column=3).font = sev_fonts.get(sev, normal_font)
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 55
            row += 1
        row += 1

    # ===== FEUILLE 2 : Inventaire =====
    ws2 = wb.create_sheet("Inventaire Hotes")
    ws2.sheet_properties.tabColor = "2E7D32"
    for col, w in {'A': 22, 'B': 16, 'C': 8, 'D': 28, 'E': 12, 'F': 16, 'G': 14}.items():
        ws2.column_dimensions[col].width = w

    ws2.merge_cells('A1:G1')
    ws2['A1'] = 'INVENTAIRE DES HOTES'
    ws2['A1'].font = title_font
    ws2.row_dimensions[1].height = 35
    for c in range(1, 8):
        ws2.cell(row=2, column=c).border = bottom_border

    inv_headers = ['Nom', 'Adresse IP', 'Agent', 'Groupes', 'Etat', 'Disponibilite', 'Type']
    for ci, hdr in enumerate(inv_headers, 1):
        cell = ws2.cell(row=3, column=ci, value=hdr)
        cell.font = header_font
        cell.fill = dark_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws2.row_dimensions[3].height = 24

    for ri, host in enumerate(hosts, 4):
        ip = at = av = ""
        if host.get("interfaces"):
            ifc = host["interfaces"][0]
            ip = ifc.get("ip", "")
            at = {"1": "ZBX", "2": "SNMP", "3": "IPMI", "4": "JMX"}.get(ifc.get("type", "1"), "?")
            av = {"0": "Inconnu", "1": "Disponible", "2": "Non disponible"}.get(ifc.get("available", "0"), "?")
        gr = ", ".join([g.get("name", "") for g in host.get("groups", [])])
        st = "Active" if str(host.get("status", "0")) == "0" else "Desactive"
        cat = classify_host(host.get("name", ""), hosts)

        vals = [host.get("name", ""), ip, at, gr, st, av, cat]
        row_fill = light_gray if (ri % 2 == 0) else white_fill
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = row_fill

        if av == "Non disponible":
            for c in range(1, 8):
                ws2.cell(row=ri, column=c).fill = PatternFill('solid', fgColor='FFEBEE')
                ws2.cell(row=ri, column=c).font = Font(name='Calibri', size=10, color='C62828')
        elif st == "Desactive":
            for c in range(1, 8):
                ws2.cell(row=ri, column=c).fill = light_gray
                ws2.cell(row=ri, column=c).font = Font(name='Calibri', size=10, color='999999')
        if cat in cat_configs:
            ws2.cell(row=ri, column=7).fill = cat_configs[cat]["fill"]
            ws2.cell(row=ri, column=7).font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
            ws2.cell(row=ri, column=7).alignment = Alignment(horizontal='center')

    # ===== FEUILLE 3 : Alertes filtrees =====
    ws3 = wb.create_sheet("Alertes filtrees")
    ws3.sheet_properties.tabColor = "999999"
    for col, w in {'A': 16, 'B': 18, 'C': 12, 'D': 55, 'E': 22}.items():
        ws3.column_dimensions[col].width = w

    ws3.merge_cells('A1:E1')
    ws3['A1'] = f'ALERTES FILTREES ({excluded_count})'
    ws3['A1'].font = Font(name='Calibri', bold=True, size=14, color='999999')

    for ci, hdr in enumerate(['Date', 'Hote', 'Severite', 'Probleme', 'Raison'], 1):
        cell = ws3.cell(row=3, column=ci, value=hdr)
        cell.font = Font(name='Calibri', bold=True, size=10, color='666666')
        cell.fill = PatternFill('solid', fgColor='EEEEEE')
        cell.border = thin_border

    frow = 4
    for p in problems:
        if not is_excluded(p):
            continue
        ti = triggers_map.get(p.get("objectid", ""), {})
        hosts_list = ti.get("hosts") or []
        hn = hosts_list[0].get("name", "") if hosts_list else ""
        try:
            et = datetime.fromtimestamp(int(p.get("clock", 0)))
        except:
            et = now
        reason = "Severite Information" if str(p.get("severity", "0")) in EXCLUDED_SEVERITIES else "Pattern exclu"
        vals = [et.strftime("%d/%m/%Y %H:%M"), hn, severity_name(p.get("severity", "0")),
                p.get("name", ""), reason]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=frow, column=ci, value=v)
            cell.font = Font(name='Calibri', size=9, color='999999')
            cell.border = thin_border
        frow += 1

    return wb


# ============================================================
# ENVOI EMAIL
# ============================================================
def send_email(filepath, filename):
    msg = MIMEMultipart()
    msg['From'] = SMTP_FROM
    msg['To'] = ", ".join(EMAIL_TO)
    msg['Subject'] = f"Rapport Zabbix hebdomadaire - {datetime.now().strftime('%d/%m/%Y')}"
    msg.attach(MIMEText(
        f"Bonjour,\n\nVeuillez trouver ci-joint le rapport Zabbix de la semaine.\n"
        f"Les recommandations sont generees par IA (Gemma3, local).\n\n"
        f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')}.\n\n---\nIT Support (IT Support)",
        'plain', 'utf-8'))

    with open(filepath, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_FROM, EMAIL_TO, msg.as_string())
    print(f"[OK] Email envoye a {', '.join(EMAIL_TO)}")


# ============================================================
# MAIN
# ============================================================
def main():
    os.makedirs(REPORT_DIR, exist_ok=True)

    if "--test-email" in sys.argv:
        print("Test d'envoi d'email...")
        try:
            msg = MIMEMultipart()
            msg['From'] = SMTP_FROM
            msg['To'] = ", ".join(EMAIL_TO)
            msg['Subject'] = "Test - Rapport Zabbix (IA)"
            msg.attach(MIMEText("Test avec recommandations IA. Configuration OK.", 'plain', 'utf-8'))
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_FROM, EMAIL_TO, msg.as_string())
            print("[OK] Email de test envoye !")
        except Exception as e:
            print(f"[ERREUR] {e}")
        return

    # Verifier Ollama
    print("Verification Ollama...", end=" ", flush=True)
    try:
        test = ask_ollama("Reponds OK")
        if test:
            print("[OK] Ollama operationnel")
        else:
            print("[WARN] Ollama ne repond pas")
    except:
        print("[WARN] Ollama indisponible")

    # Zabbix
    print("Connexion a l'API Zabbix...")
    auth = get_auth_token()
    print("[OK] Connecte")

    print("Recuperation des hotes...")
    hosts = get_hosts(auth)
    print(f"[OK] {len(hosts)} hotes")

    print("Recuperation de la disponibilite...")
    availability = get_host_availability(auth)
    print(f"[OK] {availability['available']} dispo / {availability['unavailable']} down / {availability['unknown']} inconnus")

    print("Recuperation des problemes...")
    problems = get_problems(auth)
    print(f"[OK] {len(problems)} problemes bruts")

    trigger_ids = list({p.get("objectid") for p in problems if p.get("objectid")})
    print("Recuperation des triggers...")
    triggers = get_triggers(auth, trigger_ids)
    triggers_map = {t["triggerid"]: t for t in triggers}
    print(f"[OK] {len(triggers)} triggers")

    try:
        zabbix_api("user.logout", [], auth)
    except:
        pass

    filtered = [p for p in problems if not is_excluded(p)]
    print(f"[OK] {len(filtered)} alertes pertinentes ({len(problems)-len(filtered)} filtrees)")
    print(f"Generation des recommandations IA ({len(filtered)} alertes)...")

    wb = generate_report(hosts, problems, triggers_map, availability)
    filename = f"rapport_zabbix_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)
    wb.save(filepath)
    print(f"[OK] Rapport: {filepath}")

    if "--no-email" not in sys.argv:
        print("Envoi par email...")
        try:
            send_email(filepath, filename)
        except Exception as e:
            print(f"[ERREUR] Email: {e}")

    print("\nTermine !")


if __name__ == "__main__":
    main()
